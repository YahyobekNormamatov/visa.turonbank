from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

NO_FILL = PatternFill(fill_type=None)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _med_border():
    m = Side(style="medium")
    return Border(left=m, right=m, top=m, bottom=m)

def _fmt_num(cell, val, is_int=False):
    thin = _border("thin")
    cell.border = thin
    if val is None:
        cell.value = "-"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return
    cell.value = val
    if is_int:
        cell.number_format = "#,##0"
    else:
        cell.number_format = '#,##0.00;-#,##0.00;"-"'
    cell.alignment = Alignment(horizontal="right", vertical="center")

def build_detailed_excel(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed"
    ws.freeze_panes = "A2"

    thin     = _border("thin")
    med      = _border("medium")
    bold_f   = Font(name="Calibri", bold=True, size=10)
    norm_f   = Font(name="Calibri", size=10)
    hdr_f    = Font(name="Calibri", bold=True, size=10)

    ws.column_dimensions["A"].width = 44
    for col in "BCDE":
        ws.column_dimensions[col].width = 17
    ws.row_dimensions[1].height = 4

    row_n    = [1]
    data_idx = [0]

    def _cell(col, val=None, font=None, bold=False, align="left", border=thin):
        cx = ws.cell(row=row_n[0], column=col, value=val)
        cx.font   = font or (Font(name="Calibri", bold=bold, size=10))
        cx.border = border
        cx.fill   = NO_FILL
        cx.alignment = Alignment(horizontal=align, vertical="center",
                                  wrap_text=False)
        return cx

    def next_row(h=15):
        ws.row_dimensions[row_n[0]].height = h
        row_n[0] += 1

    for r in rows:
        t = r["t"]
        data_idx[0] += 1

        if t == "banner":
            ws.merge_cells(start_row=row_n[0], start_column=1,
                           end_row=row_n[0], end_column=5)
            cx = ws.cell(row=row_n[0], column=1, value=r["text"])
            cx.font   = Font(name="Calibri", bold=True, size=12)
            cx.border = med
            cx.fill   = NO_FILL
            cx.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_n[0]].height = 22
            row_n[0] += 1

        elif t == "kv2":
            cx1 = _cell(1, f"{r['k1']}: {r['v1']}", font=bold_f, border=thin)
            if r["k2"]:
                cx2 = _cell(2, f"{r['k2']}: {r['v2']}", font=bold_f, border=thin)
                ws.merge_cells(start_row=row_n[0], start_column=2,
                               end_row=row_n[0], end_column=5)
            else:
                for c in range(2, 6):
                    _cell(c, border=thin)
            next_row(16)

        elif t == "h4":
            _cell(1, r["sec"], font=hdr_f, border=thin)
            col_names = r["cols"]
            for ci, h in enumerate(col_names, 2):
                cx = _cell(ci, h, font=hdr_f, align="center", border=thin)
            next_row(17)

        elif t == "h3opt":
            _cell(1, r["sec"], font=hdr_f, border=thin)
            _cell(2, border=thin)
            for ci, h in enumerate(r["cols"], 3):
                _cell(ci, h, font=hdr_f, align="center", border=thin)
            next_row(17)

        elif t == "h1":
            _cell(1, r["sec"], font=hdr_f, border=thin)
            _cell(2, "Amount", font=hdr_f, align="center", border=thin)
            for c in range(3, 6):
                _cell(c, border=thin)
            next_row(17)

        elif t == "d4":
            _cell(1, r["label"], font=norm_f, border=thin)
            cc = ws.cell(row=row_n[0], column=2)
            cc.font = norm_f; cc.border = thin; cc.fill = NO_FILL
            if r["count"] is not None:
                cc.value = int(r["count"])
                cc.number_format = "#,##0"
                cc.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cc.value = "-"; cc.alignment = Alignment(horizontal="center", vertical="center")
            for ci, key in enumerate(["interch", "reimb", "net"], 3):
                cx = ws.cell(row=row_n[0], column=ci)
                cx.font = norm_f; cx.fill = NO_FILL
                _fmt_num(cx, r.get(key))
            next_row(15)

        elif t == "tot4":
            _cell(1, r["label"], font=bold_f, border=thin)
            cc = ws.cell(row=row_n[0], column=2)
            cc.font = bold_f; cc.border = thin; cc.fill = NO_FILL
            if r["count"] is not None:
                cc.value = int(r["count"])
                cc.number_format = "#,##0"
                cc.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cc.value = "-"; cc.alignment = Alignment(horizontal="center", vertical="center")
            for ci, key in enumerate(["interch", "reimb", "net"], 3):
                cx = ws.cell(row=row_n[0], column=ci)
                cx.font = bold_f; cx.fill = NO_FILL
                _fmt_num(cx, r.get(key))
            next_row(16)

        elif t == "tot4plain":
            _cell(1, border=thin)
            for ci, key in enumerate(["count", "interch", "reimb", "net"], 2):
                cx = ws.cell(row=row_n[0], column=ci)
                cx.font = bold_f; cx.fill = NO_FILL
                if key == "count":
                    _fmt_num(cx, r.get(key), is_int=True)
                else:
                    _fmt_num(cx, r.get(key))
            next_row(15)

        elif t == "d3":
            _cell(1, r["label"], font=norm_f, border=thin)
            _cell(2, border=thin)
            for ci, key in enumerate(["interch", "conv", "opt"], 3):
                cx = ws.cell(row=row_n[0], column=ci)
                cx.font = norm_f; cx.fill = NO_FILL
                _fmt_num(cx, r.get(key))
            next_row(15)

        elif t == "tot3":
            _cell(1, r["label"], font=bold_f, border=thin)
            _cell(2, border=thin)
            for ci, key in enumerate(["interch", "conv", "opt"], 3):
                cx = ws.cell(row=row_n[0], column=ci)
                cx.font = bold_f; cx.fill = NO_FILL
                _fmt_num(cx, r.get(key))
            next_row(16)

        elif t == "d1":
            _cell(1, r["label"], font=norm_f, border=thin)
            cx = ws.cell(row=row_n[0], column=2)
            cx.font = norm_f; cx.fill = NO_FILL
            _fmt_num(cx, r.get("val"))
            for c in range(3, 6):
                _cell(c, border=thin)
            next_row(15)

        elif t == "tot1":
            _cell(1, r["label"], font=bold_f, border=thin)
            cx = ws.cell(row=row_n[0], column=2)
            cx.font = bold_f; cx.fill = NO_FILL
            _fmt_num(cx, r.get("val"))
            for c in range(3, 6):
                _cell(c, border=thin)
            next_row(16)

        elif t == "text":
            _cell(1, r["label"], font=norm_f, border=thin)
            for c in range(2, 6):
                _cell(c, border=thin)
            next_row(15)

    wb.save(out_path)


SUMMARY_ROWS = [
    ("Purchase", "Issuer", "Purchase"),

    ("ATM Cash", "Issuer", "ATM Cash"),

    ("Manual Cash", "Issuer", "Manual Cash"),

    ("Quasi-cash", "Issuer", "Quasi-cash"),

    ("Quasi-cash Credit", "Issuer", "Quasi-cash Credit"),

    ("Cardholder Funds Transfer", "Issuer", "Cardholder Funds Transfer"),

    ("Merchandise Credit", "Issuer", "Merchandise Credit"),

    ("Manual Cash", "Acquirer", "Manual Cash экваир"),

    ("ATM Balance Inquiry","Issuer", "ATM Balance Inquiry"),

    ("ATM Decline", "Issuer", "ATM Decline"),

    ("FC - V.I.P. System", "Issuer", "FC - V.I.P. System"),

    ("ISA CHARGE MULTI CURR CASH", "Issuer", "ISA CHARGE MULTI CURR CASH"),

    ("ISA CHARGE MULTI CURR PUR", "Issuer", "ISA CHARGE MULTI CURR PUR"),

    ("ISA CHARGE SINGLE CURR CASH", "Issuer", "ISA CHARGE SINGLE CURR CASH"),

    ("ISA CHARGE SINGLE CURR PUR", "Issuer", "ISA CHARGE SINGLE CURR PUR"),

    ("IAF CHARGE ENHANCED CASH", "Issuer", "IAF CHARGE ENHANCED CASH"),
]

ISA_IAF_LABELS = {
    "ISA CHARGE MULTI CURR CASH", "ISA CHARGE MULTI CURR PUR",
    "ISA CHARGE SINGLE CURR CASH", "ISA CHARGE SINGLE CURR PUR",
    "IAF CHARGE ENHANCED CASH",
}


def _aggregate(rows, group_name):
    agg = defaultdict(lambda: {"interch": 0.0, "reimb": 0.0})
    cur_group   = None
    cur_section = None
    for r in rows:
        if r["t"] == "kv2" and r["k1"] == "Reporting for":
            cur_group = r["v1"]
        if r["t"] == "h4":
            cur_section = r["sec"]
        if cur_group != group_name:
            continue
        if r["t"] == "d4":
            side = "Acquirer" if cur_section and "Acquirer" in cur_section else "Issuer"
            key  = (r["label"], side)
            agg[key]["interch"] += r.get("interch") or 0.0
            agg[key]["reimb"]   += r.get("reimb")   or 0.0
        if r["t"] == "d1":
            key = (r["label"], "Issuer")
            agg[key]["interch"] += r.get("val") or 0.0
    return agg


def _write_summary_sheet(ws, rows, group_name, title_label, is_uzs=False):
    thin   = _border("thin")
    bold_f = Font(name="Calibri", bold=True, size=10)
    norm_f = Font(name="Calibri", size=10)
    num_fmt = '#,##0.00;-#,##0.00;"-"' if not is_uzs else '#,##0;-#,##0;"-"'

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22

    def sc(r, c, val=None, font=None, align="left"):
        cx = ws.cell(row=r, column=c, value=val)
        cx.font   = font or norm_f
        cx.border = thin
        cx.fill   = NO_FILL
        cx.alignment = Alignment(horizontal=align, vertical="center")
        return cx

    def nc(r, c, val, font=None):
        cx = ws.cell(row=r, column=c)
        cx.font   = font or norm_f
        cx.border = thin
        cx.fill   = NO_FILL
        if val is None or val == 0.0:
            cx.value = "-"
            cx.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cx.value = val
            cx.number_format = num_fmt
            cx.alignment = Alignment(horizontal="right", vertical="center")
        return cx

    ws.merge_cells("A1:F1")
    c1 = ws.cell(row=1, column=1, value=title_label)
    c1.font  = Font(name="Calibri", bold=True, size=13)
    c1.border = thin; c1.fill = NO_FILL
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:B2")
    for c in (1, 2):
        cx = ws.cell(row=2, column=c); cx.border = thin; cx.fill = NO_FILL
    ws.merge_cells("C2:D2")
    hs = ws.cell(row=2, column=3, value="Сумма")
    hs.font = bold_f; hs.border = thin; hs.fill = NO_FILL
    hs.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=4).border = thin; ws.cell(row=2, column=4).fill = NO_FILL
    ws.merge_cells("E2:F2")
    hk = ws.cell(row=2, column=5, value="Комиссия")
    hk.font = bold_f; hk.border = thin; hk.fill = NO_FILL
    hk.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=6).border = thin; ws.cell(row=2, column=6).fill = NO_FILL
    ws.row_dimensions[2].height = 18

    sub_hdrs = ["№", title_label, "Debit", "Credit", "Reim fee Debit", "Interchange fee Credit"]
    for ci, h in enumerate(sub_hdrs, 1):
        cx = ws.cell(row=3, column=ci, value=h)
        cx.font = bold_f; cx.border = thin; cx.fill = NO_FILL
        cx.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    agg = _aggregate(rows, group_name)
    displayed = []
    for label_key, side, display_name in SUMMARY_ROWS:
        v = agg.get((label_key, side), {"interch": 0.0, "reimb": 0.0})
        interch = v["interch"]; reimb = v["reimb"]
        if label_key in ISA_IAF_LABELS:
            debit = credit = interch_cred = None
            reim_debit = abs(interch) if interch < 0 else None
        else:
            debit        = abs(interch) if interch < 0 else None
            credit       = interch      if interch > 0 else None
            reim_debit   = abs(reimb)   if reimb   < 0 else None
            interch_cred = reimb        if reimb   > 0 else None
        if all(x is None for x in [debit, credit, reim_debit, interch_cred]):
            continue
        displayed.append((display_name, debit, credit, reim_debit, interch_cred))

    data_start = 4
    for i, (name, debit, credit, reim_d, ic) in enumerate(displayed):
        r = data_start + i
        sc(r, 1, i + 1, align="center")
        sc(r, 2, name)
        nc(r, 3, debit); nc(r, 4, credit)
        nc(r, 5, reim_d); nc(r, 6, ic)
        ws.row_dimensions[r].height = 17

    total_r = data_start + len(displayed)
    ws.merge_cells(f"A{total_r}:B{total_r}")
    itogo = ws.cell(row=total_r, column=1, value="Итого")
    itogo.font = bold_f; itogo.border = thin; itogo.fill = NO_FILL
    itogo.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_r, column=2).border = thin
    ws.cell(row=total_r, column=2).fill   = NO_FILL
    for ci in range(3, 7):
        col_letter = get_column_letter(ci)
        cx = ws.cell(row=total_r, column=ci,
                     value=f"=SUM({col_letter}{data_start}:{col_letter}{total_r-1})")
        cx.font = bold_f; cx.border = thin; cx.fill = NO_FILL
        cx.number_format = num_fmt
        cx.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[total_r].height = 19

    netto_r = total_r + 1
    ws.merge_cells(f"A{netto_r}:B{netto_r}")
    nl = ws.cell(row=netto_r, column=1, value="Netto value")
    nl.font = bold_f; nl.border = thin; nl.fill = NO_FILL
    nl.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=netto_r, column=2).border = thin
    ws.cell(row=netto_r, column=2).fill   = NO_FILL
    nv = ws.cell(row=netto_r, column=3,
                 value=f"=(D{total_r}+F{total_r})-(C{total_r}+E{total_r})")
    nv.font = bold_f; nv.border = thin; nv.fill = NO_FILL
    nv.number_format = num_fmt
    nv.alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"C{netto_r}:F{netto_r}")
    for ci in range(4, 7):
        cx = ws.cell(row=netto_r, column=ci)
        cx.border = thin; cx.fill = NO_FILL
    ws.row_dimensions[netto_r].height = 19


def build_all_summaries_excel(rows, out_path):
    SHEET_SPECS = [
        ("TURON HUMO USD", "HUMO USD", False),
        ("TURON SUMMARY USD", "SUMMARY USD", False),
        ("TURON HUMO UZS", "HUMO UZS", True),
        ("TURON SUMMARY UZS", "SUMMARY UZS", True),
    ]
    wb = Workbook()
    wb.remove(wb.active)
    for group_name, sheet_title, is_uzs in SHEET_SPECS:
        ws = wb.create_sheet(title=sheet_title)
        _write_summary_sheet(ws, rows, group_name, group_name, is_uzs=is_uzs)
    wb.save(out_path)

def build_summary_excel(rows, group_name, title_label, out_path, is_uzs=False):
    wb = Workbook()
    ws = wb.active
    ws.title = title_label[:31]
    _write_summary_sheet(ws, rows, group_name, title_label, is_uzs=is_uzs)
    wb.save(out_path)